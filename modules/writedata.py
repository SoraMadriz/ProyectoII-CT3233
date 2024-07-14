import openpyxl as op

def writenResults(array,col,rows,salida,sheet):
    path = f'./files/{salida}.xlsx'
    print(path)
    wb = op.load_workbook(path)
    sheet = wb[f'{sheet}']

    if len(array) > 1:
        for i in range(len(array)):
            sheet.cell(row=i+rows,column=col).value = array[i]
    elif len(array) == 1:
        sheet.cell(row=rows,column=col).value = array[0]

    wb.save(path)
